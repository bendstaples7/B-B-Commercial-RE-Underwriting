"""Report Generator service for creating comprehensive analysis reports."""
from typing import Dict, List, Optional, Any
from datetime import datetime
from app.models.analysis_session import AnalysisSession
from app.models.property_facts import PropertyFacts, PropertyType
from app.models.comparable_sale import ComparableSale
from app.models.ranked_comparable import RankedComparable
from app.models.valuation_result import ValuationResult, ComparableValuation
from app.models.scenario import Scenario, WholesaleScenario, FixFlipScenario, BuyHoldScenario


class ReportGenerator:
    """Generates comprehensive analysis reports with export capabilities."""
    
    def _is_residential(self, property_type: PropertyType) -> bool:
        """
        Check if property type is residential.
        
        Args:
            property_type: Property type
            
        Returns:
            True if residential, False otherwise
        """
        return property_type in [PropertyType.SINGLE_FAMILY, PropertyType.MULTI_FAMILY]
    
    def _get_valuation_method_label(self, property_type: PropertyType, method: str) -> str:
        """
        Get property-type-specific label for valuation method.
        
        Args:
            property_type: Property type
            method: Valuation method name
            
        Returns:
            Formatted label for the method
        """
        is_residential = self._is_residential(property_type)
        
        labels = {
            'price_per_sqft': 'Price per Sq Ft',
            'price_per_unit': 'Price per Unit',
            'price_per_bedroom': 'Price per Bedroom' if is_residential else 'Income Capitalization',
            'adjusted_value': 'Adjusted Value'
        }
        
        return labels.get(method, method.replace('_', ' ').title())
    
    def generate_report(self, session: AnalysisSession) -> Dict[str, Any]:
        """
        Generate complete report with all sections.
        
        Args:
            session: AnalysisSession with complete workflow data
            
        Returns:
            Dictionary containing all report sections
        """
        report = {
            'session_id': session.session_id,
            'generated_at': datetime.utcnow().isoformat(),
            'sections': {}
        }
        
        # Get property type for terminology
        property_type = session.subject_property.property_type if session.subject_property else None
        
        # Section A: Subject Property Facts
        if session.subject_property:
            report['sections']['section_a'] = self.format_section_a(session.subject_property)
        
        # Section B: Comparable Sales
        comparables = session.comparables.all()
        if comparables:
            report['sections']['section_b'] = self.format_section_b(
                session.subject_property, 
                comparables
            )
        
        # Section C: Weighted Ranking
        ranked_comparables = session.ranked_comparables.order_by(RankedComparable.rank).all()
        if ranked_comparables:
            report['sections']['section_c'] = self.format_section_c(ranked_comparables)
        
        # Section D: Valuation Models (with property-type-specific terminology)
        if session.valuation_result:
            report['sections']['section_d'] = self.format_section_d(
                session.valuation_result,
                property_type
            )
        
        # Section E: ARV Range
        if session.valuation_result:
            report['sections']['section_e'] = self.format_section_e(session.valuation_result)
        
        # Section F: Key Drivers
        if session.valuation_result and session.valuation_result.key_drivers:
            report['sections']['section_f'] = self.format_section_f(
                session.valuation_result.key_drivers
            )
        
        # Optional: Scenario Analysis
        scenarios = session.scenarios.all()
        if scenarios:
            report['sections']['scenarios'] = self._format_scenario_sections(scenarios)
        
        return report
    
    def format_section_a(self, subject: PropertyFacts) -> Dict[str, Any]:
        """
        Format Section A: Subject Property Facts table.
        Uses property-type-specific terminology.
        
        Args:
            subject: PropertyFacts object
            
        Returns:
            Dictionary with property facts data
        """
        is_residential = self._is_residential(subject.property_type)
        
        data = {
            'Address': subject.address,
            'Property Type': subject.property_type.value.replace('_', ' ').title(),
            'Units': subject.units,
        }
        
        # Add residential-specific fields
        if is_residential:
            data['Bedrooms'] = subject.bedrooms
            data['Bathrooms'] = subject.bathrooms
        
        # Common fields
        data.update({
            'Square Footage': f"{subject.square_footage:,}",
            'Lot Size': f"{subject.lot_size:,} sq ft",
            'Year Built': subject.year_built,
            'Construction Type': subject.construction_type.value.title(),
        })
        
        # Add basement for residential only
        if is_residential:
            data['Basement'] = 'Yes' if subject.basement else 'No'
        
        data.update({
            'Parking Spaces': subject.parking_spaces,
            'Last Sale Price': f"${subject.last_sale_price:,.0f}" if subject.last_sale_price else 'N/A',
            'Last Sale Date': subject.last_sale_date.strftime('%m/%d/%Y') if subject.last_sale_date else 'N/A',
            'Assessed Value': f"${subject.assessed_value:,.0f}",
            'Annual Taxes': f"${subject.annual_taxes:,.0f}",
            'Zoning': subject.zoning,
            'Interior Condition': subject.interior_condition.value.replace('_', ' ').title(),
            'Data Source': subject.data_source or 'Multiple Sources',
            'User Modified Fields': ', '.join(subject.user_modified_fields) if subject.user_modified_fields else 'None'
        })
        
        return {
            'title': 'Section A: Subject Property Facts',
            'data': data,
            'property_type': subject.property_type.value
        }

    def format_section_b(self, subject: PropertyFacts, comparables: List[ComparableSale]) -> Dict[str, Any]:
        """
        Format Section B: Comparable Sales table.
        
        Args:
            subject: Subject property (displayed as first row)
            comparables: List of comparable sales
            
        Returns:
            Dictionary with comparable sales data
        """
        rows = []
        
        # First row: Subject property
        subject_row = {
            'Address': subject.address,
            'Sale Date': subject.last_sale_date.strftime('%m/%d/%Y') if subject.last_sale_date else 'N/A',
            'Sale Price': f"${subject.last_sale_price:,.0f}" if subject.last_sale_price else 'N/A',
            'Property Type': subject.property_type.value.replace('_', ' ').title(),
            'Units': subject.units,
            'Bedrooms': subject.bedrooms,
            'Bathrooms': subject.bathrooms,
            'Square Footage': f"{subject.square_footage:,}",
            'Lot Size': f"{subject.lot_size:,}",
            'Year Built': subject.year_built,
            'Construction': subject.construction_type.value.title(),
            'Interior': subject.interior_condition.value.replace('_', ' ').title(),
            'Distance': 'Subject',
            'Type': 'Subject Property'
        }
        rows.append(subject_row)
        
        # Subsequent rows: Comparables
        for comp in comparables:
            comp_row = {
                'Address': comp.address,
                'Sale Date': comp.sale_date.strftime('%m/%d/%Y'),
                'Sale Price': f"${comp.sale_price:,.0f}",
                'Property Type': comp.property_type.value.replace('_', ' ').title(),
                'Units': comp.units,
                'Bedrooms': comp.bedrooms,
                'Bathrooms': comp.bathrooms,
                'Square Footage': f"{comp.square_footage:,}",
                'Lot Size': f"{comp.lot_size:,}",
                'Year Built': comp.year_built,
                'Construction': comp.construction_type.value.title(),
                'Interior': comp.interior_condition.value.replace('_', ' ').title(),
                'Distance': f"{comp.distance_miles:.2f} mi",
                'Type': 'Comparable'
            }
            rows.append(comp_row)
        
        return {
            'title': 'Section B: Comparable Sales',
            'columns': [
                'Address', 'Sale Date', 'Sale Price', 'Property Type', 'Units',
                'Bedrooms', 'Bathrooms', 'Square Footage', 'Lot Size', 'Year Built',
                'Construction', 'Interior', 'Distance', 'Type'
            ],
            'rows': rows
        }
    
    def format_section_c(self, ranked_comparables: List[RankedComparable]) -> Dict[str, Any]:
        """
        Format Section C: Weighted Ranking table.
        
        Args:
            ranked_comparables: List of ranked comparables with scores
            
        Returns:
            Dictionary with weighted ranking data
        """
        rows = []
        
        for ranked in ranked_comparables:
            comp = ranked.comparable
            row = {
                'Rank': ranked.rank,
                'Address': comp.address,
                'Recency (16%)': f"{ranked.recency_score:.1f}",
                'Proximity (15%)': f"{ranked.proximity_score:.1f}",
                'Units (15%)': f"{ranked.units_score:.1f}",
                'Beds/Baths (15%)': f"{ranked.beds_baths_score:.1f}",
                'Sq Ft (15%)': f"{ranked.sqft_score:.1f}",
                'Construction (12%)': f"{ranked.construction_score:.1f}",
                'Interior (12%)': f"{ranked.interior_score:.1f}",
                'Total Score': f"{ranked.total_score:.2f}"
            }
            rows.append(row)
        
        return {
            'title': 'Section C: Weighted Ranking',
            'columns': [
                'Rank', 'Address', 'Recency (16%)', 'Proximity (15%)', 'Units (15%)',
                'Beds/Baths (15%)', 'Sq Ft (15%)', 'Construction (12%)', 'Interior (12%)',
                'Total Score'
            ],
            'rows': rows
        }
    
    def format_section_d(self, valuation_result: ValuationResult, property_type: PropertyType = None) -> Dict[str, Any]:
        """
        Format Section D: Valuation Models with narratives.
        Uses property-type-specific terminology for valuation methods.
        
        Args:
            valuation_result: ValuationResult with comparable valuations
            property_type: Property type for terminology selection
            
        Returns:
            Dictionary with valuation models data
        """
        valuations = []
        
        # Determine property type from first comparable if not provided
        if property_type is None and valuation_result.comparable_valuations:
            # Try to get from session or default to residential
            property_type = PropertyType.SINGLE_FAMILY
        
        is_residential = self._is_residential(property_type)
        
        for comp_val in valuation_result.comparable_valuations:
            comp = comp_val.comparable
            
            # Build metrics with property-type-specific labels
            metrics = {
                'Price per Sq Ft': f"${comp_val.price_per_sqft:,.2f}",
                'Price per Unit': f"${comp_val.price_per_unit:,.0f}",
                self._get_valuation_method_label(property_type, 'price_per_bedroom'): f"${comp_val.price_per_bedroom:,.0f}",
                'Adjusted Value': f"${comp_val.adjusted_value:,.0f}"
            }
            
            valuation = {
                'address': comp.address,
                'narrative': comp_val.narrative or 'No narrative available',
                'metrics': metrics,
                'adjustments': comp_val.adjustments or []
            }
            valuations.append(valuation)
        
        return {
            'title': 'Section D: Valuation Models',
            'valuations': valuations,
            'property_type': property_type.value if property_type else 'residential'
        }
    
    def format_section_e(self, valuation_result: ValuationResult) -> Dict[str, Any]:
        """
        Format Section E: ARV Range display.
        
        Args:
            valuation_result: ValuationResult with ARV range
            
        Returns:
            Dictionary with ARV range data
        """
        return {
            'title': 'Section E: Final ARV Range',
            'arv_range': {
                'Conservative (25th Percentile)': f"${valuation_result.conservative_arv:,.0f}",
                'Likely (Median)': f"${valuation_result.likely_arv:,.0f}",
                'Aggressive (75th Percentile)': f"${valuation_result.aggressive_arv:,.0f}"
            },
            'all_valuations': [f"${val:,.0f}" for val in valuation_result.all_valuations]
        }
    
    def format_section_f(self, key_drivers: List[str]) -> Dict[str, Any]:
        """
        Format Section F: Key Drivers bullet points.
        
        Args:
            key_drivers: List of key driver strings
            
        Returns:
            Dictionary with key drivers data
        """
        return {
            'title': 'Section F: Key Drivers',
            'drivers': key_drivers
        }
    
    def _format_scenario_sections(self, scenarios: List[Scenario]) -> Dict[str, Any]:
        """
        Format optional scenario analysis sections.
        
        Args:
            scenarios: List of scenario objects
            
        Returns:
            Dictionary with scenario analysis data
        """
        scenario_data = {
            'wholesale': [],
            'fix_flip': [],
            'buy_hold': []
        }
        
        for scenario in scenarios:
            if isinstance(scenario, WholesaleScenario):
                scenario_data['wholesale'].append(self._format_wholesale_scenario(scenario))
            elif isinstance(scenario, FixFlipScenario):
                scenario_data['fix_flip'].append(self._format_fix_flip_scenario(scenario))
            elif isinstance(scenario, BuyHoldScenario):
                scenario_data['buy_hold'].append(self._format_buy_hold_scenario(scenario))
        
        return scenario_data
    
    def _format_wholesale_scenario(self, scenario: WholesaleScenario) -> Dict[str, Any]:
        """Format wholesale scenario data."""
        return {
            'purchase_price': f"${scenario.purchase_price:,.0f}",
            'mao': f"${scenario.mao:,.0f}",
            'contract_price': f"${scenario.contract_price:,.0f}",
            'assignment_fee_range': f"${scenario.assignment_fee_low:,.0f} - ${scenario.assignment_fee_high:,.0f}",
            'estimated_repairs': f"${scenario.estimated_repairs:,.0f}",
            'summary': scenario.summary
        }
    
    def _format_fix_flip_scenario(self, scenario: FixFlipScenario) -> Dict[str, Any]:
        """Format fix and flip scenario data."""
        return {
            'acquisition_cost': f"${scenario.acquisition_cost:,.0f}",
            'renovation_cost': f"${scenario.renovation_cost:,.0f}",
            'holding_costs': f"${scenario.holding_costs:,.0f}",
            'financing_costs': f"${scenario.financing_costs:,.0f}",
            'closing_costs': f"${scenario.closing_costs:,.0f}",
            'total_cost': f"${scenario.total_cost:,.0f}",
            'exit_value': f"${scenario.exit_value:,.0f}",
            'net_profit': f"${scenario.net_profit:,.0f}",
            'roi': f"{scenario.roi:.2f}%",
            'months_to_flip': scenario.months_to_flip,
            'summary': scenario.summary
        }
    
    def _format_buy_hold_scenario(self, scenario: BuyHoldScenario) -> Dict[str, Any]:
        """Format buy and hold scenario data."""
        return {
            'market_rent': f"${scenario.market_rent:,.0f}",
            'capital_structures': scenario.capital_structures,
            'price_points': scenario.price_points,
            'summary': scenario.summary
        }

    def export_to_excel(self, report: Dict[str, Any]) -> bytes:
        """
        Export report to Excel format with styling.
        
        Args:
            report: Report dictionary from generate_report()
            
        Returns:
            Excel file as bytes
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        sections = report.get('sections', {})
        
        # Create Section A: Property Facts
        if 'section_a' in sections:
            ws = wb.create_sheet('Property Facts')
            self._write_property_facts_sheet(ws, sections['section_a'])
        
        # Create Section B: Comparable Sales
        if 'section_b' in sections:
            ws = wb.create_sheet('Comparable Sales')
            self._write_comparable_sales_sheet(ws, sections['section_b'])
        
        # Create Section C: Weighted Ranking
        if 'section_c' in sections:
            ws = wb.create_sheet('Weighted Ranking')
            self._write_weighted_ranking_sheet(ws, sections['section_c'])
        
        # Create Section D: Valuation Models
        if 'section_d' in sections:
            ws = wb.create_sheet('Valuation Models')
            self._write_valuation_models_sheet(ws, sections['section_d'])
        
        # Create Section E: ARV Range
        if 'section_e' in sections:
            ws = wb.create_sheet('ARV Range')
            self._write_arv_range_sheet(ws, sections['section_e'])
        
        # Create Section F: Key Drivers
        if 'section_f' in sections:
            ws = wb.create_sheet('Key Drivers')
            self._write_key_drivers_sheet(ws, sections['section_f'])
        
        # Create Scenario Analysis sheets if present
        if 'scenarios' in sections:
            self._write_scenario_sheets(wb, sections['scenarios'])
        
        # Save to bytes
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.getvalue()
    
    def _write_property_facts_sheet(self, ws, section_data: Dict[str, Any]):
        """Write property facts to Excel sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Title
        ws['A1'] = section_data['title']
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write data
        row = 3
        for key, value in section_data['data'].items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _write_comparable_sales_sheet(self, ws, section_data: Dict[str, Any]):
        """Write comparable sales to Excel sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Title
        ws['A1'] = section_data['title']
        ws['A1'].font = Font(bold=True, size=14)
        
        # Headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write column headers
        for col_idx, col_name in enumerate(section_data['columns'], start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_idx, row_data in enumerate(section_data['rows'], start=4):
            for col_idx, col_name in enumerate(section_data['columns'], start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                cell.border = border
                
                # Highlight subject property
                if row_data.get('Type') == 'Subject Property':
                    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        # Adjust column widths
        for col_idx in range(1, len(section_data['columns']) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15
    
    def _write_weighted_ranking_sheet(self, ws, section_data: Dict[str, Any]):
        """Write weighted ranking to Excel sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Title
        ws['A1'] = section_data['title']
        ws['A1'].font = Font(bold=True, size=14)
        
        # Headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write column headers
        for col_idx, col_name in enumerate(section_data['columns'], start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_idx, row_data in enumerate(section_data['rows'], start=4):
            for col_idx, col_name in enumerate(section_data['columns'], start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                cell.border = border
                
                # Highlight top 5 comparables
                if row_data.get('Rank', 999) <= 5:
                    cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        
        # Adjust column widths
        for col_idx in range(1, len(section_data['columns']) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15
    
    def _write_valuation_models_sheet(self, ws, section_data: Dict[str, Any]):
        """Write valuation models to Excel sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Title
        ws['A1'] = section_data['title']
        ws['A1'].font = Font(bold=True, size=14)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 3
        for valuation in section_data['valuations']:
            # Address header
            ws[f'A{row}'] = valuation['address']
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Narrative
            ws[f'A{row}'] = 'Narrative:'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = valuation['narrative']
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
            
            # Metrics
            ws[f'A{row}'] = 'Valuation Metrics:'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for metric_name, metric_value in valuation['metrics'].items():
                ws[f'A{row}'] = metric_name
                ws[f'B{row}'] = metric_value
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
            
            row += 2  # Spacing between comparables
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _write_arv_range_sheet(self, ws, section_data: Dict[str, Any]):
        """Write ARV range to Excel sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Title
        ws['A1'] = section_data['title']
        ws['A1'].font = Font(bold=True, size=14)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ARV Range
        row = 3
        for arv_type, arv_value in section_data['arv_range'].items():
            ws[f'A{row}'] = arv_type
            ws[f'B{row}'] = arv_value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _write_key_drivers_sheet(self, ws, section_data: Dict[str, Any]):
        """Write key drivers to Excel sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Title
        ws['A1'] = section_data['title']
        ws['A1'].font = Font(bold=True, size=14)
        
        # Drivers
        row = 3
        for idx, driver in enumerate(section_data['drivers'], start=1):
            ws[f'A{row}'] = f"{idx}. {driver}"
            row += 1
        
        # Adjust column width
        ws.column_dimensions['A'].width = 80
    
    def _write_scenario_sheets(self, wb, scenarios_data: Dict[str, Any]):
        """Write scenario analysis sheets."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Wholesale scenarios
        if scenarios_data.get('wholesale'):
            ws = wb.create_sheet('Wholesale Analysis')
            self._write_wholesale_sheet(ws, scenarios_data['wholesale'])
        
        # Fix & Flip scenarios
        if scenarios_data.get('fix_flip'):
            ws = wb.create_sheet('Fix & Flip Analysis')
            self._write_fix_flip_sheet(ws, scenarios_data['fix_flip'])
        
        # Buy & Hold scenarios
        if scenarios_data.get('buy_hold'):
            ws = wb.create_sheet('Buy & Hold Analysis')
            self._write_buy_hold_sheet(ws, scenarios_data['buy_hold'])
    
    def _write_wholesale_sheet(self, ws, scenarios: List[Dict[str, Any]]):
        """Write wholesale scenarios to sheet."""
        from openpyxl.styles import Font, Border, Side
        
        ws['A1'] = 'Wholesale Analysis'
        ws['A1'].font = Font(bold=True, size=14)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 3
        for scenario in scenarios:
            ws[f'A{row}'] = 'Purchase Price'
            ws[f'B{row}'] = scenario['purchase_price']
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = 'MAO'
            ws[f'B{row}'] = scenario['mao']
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = 'Contract Price'
            ws[f'B{row}'] = scenario['contract_price']
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = 'Assignment Fee Range'
            ws[f'B{row}'] = scenario['assignment_fee_range']
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            ws[f'A{row}'] = 'Estimated Repairs'
            ws[f'B{row}'] = scenario['estimated_repairs']
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 3
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _write_fix_flip_sheet(self, ws, scenarios: List[Dict[str, Any]]):
        """Write fix & flip scenarios to sheet."""
        from openpyxl.styles import Font, Border, Side
        
        ws['A1'] = 'Fix & Flip Analysis'
        ws['A1'].font = Font(bold=True, size=14)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 3
        for scenario in scenarios:
            fields = [
                ('Acquisition Cost', 'acquisition_cost'),
                ('Renovation Cost', 'renovation_cost'),
                ('Holding Costs', 'holding_costs'),
                ('Financing Costs', 'financing_costs'),
                ('Closing Costs', 'closing_costs'),
                ('Total Cost', 'total_cost'),
                ('Exit Value', 'exit_value'),
                ('Net Profit', 'net_profit'),
                ('ROI', 'roi'),
                ('Months to Flip', 'months_to_flip')
            ]
            
            for label, key in fields:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = scenario[key]
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
            
            row += 2
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _write_buy_hold_sheet(self, ws, scenarios: List[Dict[str, Any]]):
        """Write buy & hold scenarios to sheet."""
        from openpyxl.styles import Font, Border, Side
        
        ws['A1'] = 'Buy & Hold Analysis'
        ws['A1'].font = Font(bold=True, size=14)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 3
        for scenario in scenarios:
            ws[f'A{row}'] = 'Market Rent'
            ws[f'B{row}'] = scenario['market_rent']
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 2
            
            # Note: Capital structures and price points are complex nested data
            # For simplicity, we'll just note they exist
            ws[f'A{row}'] = 'See summary data for detailed capital structures and price points'
            row += 2
        
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 20

    def export_to_google_sheets(self, report: Dict[str, Any], user_credentials: Dict[str, Any]) -> str:
        """
        Export report to Google Sheets with OAuth authentication.
        
        Args:
            report: Report dictionary from generate_report()
            user_credentials: OAuth credentials dictionary
            
        Returns:
            Shareable URL of the created Google Sheet
        """
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        from googleapiclient.errors import HttpError
        
        try:
            # Create credentials from user credentials dict
            creds = Credentials(
                token=user_credentials.get('token'),
                refresh_token=user_credentials.get('refresh_token'),
                token_uri=user_credentials.get('token_uri', 'https://oauth2.googleapis.com/token'),
                client_id=user_credentials.get('client_id'),
                client_secret=user_credentials.get('client_secret'),
                scopes=user_credentials.get('scopes', ['https://www.googleapis.com/auth/spreadsheets'])
            )
            
            # Build the Sheets API service
            service = build('sheets', 'v4', credentials=creds)
            
            # Create a new spreadsheet
            spreadsheet_title = f"Real Estate Analysis - {report.get('session_id', 'Report')} - {datetime.utcnow().strftime('%Y-%m-%d')}"
            spreadsheet = {
                'properties': {
                    'title': spreadsheet_title
                }
            }
            
            spreadsheet = service.spreadsheets().create(body=spreadsheet).execute()
            spreadsheet_id = spreadsheet['spreadsheetId']
            
            # Prepare batch update requests
            requests = []
            sections = report.get('sections', {})
            
            # Add sheets for each section
            sheet_id = 0
            
            # Section A: Property Facts
            if 'section_a' in sections:
                requests.extend(self._create_property_facts_requests(sections['section_a'], sheet_id))
                sheet_id += 1
            
            # Section B: Comparable Sales
            if 'section_b' in sections:
                requests.extend(self._create_comparable_sales_requests(sections['section_b'], sheet_id))
                sheet_id += 1
            
            # Section C: Weighted Ranking
            if 'section_c' in sections:
                requests.extend(self._create_weighted_ranking_requests(sections['section_c'], sheet_id))
                sheet_id += 1
            
            # Section D: Valuation Models
            if 'section_d' in sections:
                requests.extend(self._create_valuation_models_requests(sections['section_d'], sheet_id))
                sheet_id += 1
            
            # Section E: ARV Range
            if 'section_e' in sections:
                requests.extend(self._create_arv_range_requests(sections['section_e'], sheet_id))
                sheet_id += 1
            
            # Section F: Key Drivers
            if 'section_f' in sections:
                requests.extend(self._create_key_drivers_requests(sections['section_f'], sheet_id))
                sheet_id += 1
            
            # Execute batch update
            if requests:
                body = {'requests': requests}
                service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body=body
                ).execute()
            
            # Return shareable URL
            return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
            
        except HttpError as error:
            raise Exception(f"Google Sheets API error: {error}")
        except Exception as error:
            raise Exception(f"Failed to export to Google Sheets: {error}")
    
    def _create_property_facts_requests(self, section_data: Dict[str, Any], sheet_id: int) -> List[Dict[str, Any]]:
        """Create Google Sheets requests for property facts section."""
        requests = []
        
        # Add sheet
        requests.append({
            'addSheet': {
                'properties': {
                    'sheetId': sheet_id,
                    'title': 'Property Facts'
                }
            }
        })
        
        # Prepare data rows
        rows = []
        rows.append([section_data['title'], ''])  # Title row
        rows.append(['', ''])  # Empty row
        
        for key, value in section_data['data'].items():
            rows.append([key, str(value)])
        
        # Update cells
        requests.append({
            'updateCells': {
                'rows': [{'values': [{'userEnteredValue': {'stringValue': str(cell)}} for cell in row]} for row in rows],
                'fields': 'userEnteredValue',
                'start': {'sheetId': sheet_id, 'rowIndex': 0, 'columnIndex': 0}
            }
        })
        
        return requests
    
    def _create_comparable_sales_requests(self, section_data: Dict[str, Any], sheet_id: int) -> List[Dict[str, Any]]:
        """Create Google Sheets requests for comparable sales section."""
        requests = []
        
        # Add sheet
        requests.append({
            'addSheet': {
                'properties': {
                    'sheetId': sheet_id,
                    'title': 'Comparable Sales'
                }
            }
        })
        
        # Prepare data rows
        rows = []
        rows.append([section_data['title']])  # Title row
        rows.append([])  # Empty row
        rows.append(section_data['columns'])  # Header row
        
        for row_data in section_data['rows']:
            row = [str(row_data.get(col, '')) for col in section_data['columns']]
            rows.append(row)
        
        # Update cells
        requests.append({
            'updateCells': {
                'rows': [{'values': [{'userEnteredValue': {'stringValue': str(cell)}} for cell in row]} for row in rows],
                'fields': 'userEnteredValue',
                'start': {'sheetId': sheet_id, 'rowIndex': 0, 'columnIndex': 0}
            }
        })
        
        return requests
    
    def _create_weighted_ranking_requests(self, section_data: Dict[str, Any], sheet_id: int) -> List[Dict[str, Any]]:
        """Create Google Sheets requests for weighted ranking section."""
        requests = []
        
        # Add sheet
        requests.append({
            'addSheet': {
                'properties': {
                    'sheetId': sheet_id,
                    'title': 'Weighted Ranking'
                }
            }
        })
        
        # Prepare data rows
        rows = []
        rows.append([section_data['title']])  # Title row
        rows.append([])  # Empty row
        rows.append(section_data['columns'])  # Header row
        
        for row_data in section_data['rows']:
            row = [str(row_data.get(col, '')) for col in section_data['columns']]
            rows.append(row)
        
        # Update cells
        requests.append({
            'updateCells': {
                'rows': [{'values': [{'userEnteredValue': {'stringValue': str(cell)}} for cell in row]} for row in rows],
                'fields': 'userEnteredValue',
                'start': {'sheetId': sheet_id, 'rowIndex': 0, 'columnIndex': 0}
            }
        })
        
        return requests
    
    def _create_valuation_models_requests(self, section_data: Dict[str, Any], sheet_id: int) -> List[Dict[str, Any]]:
        """Create Google Sheets requests for valuation models section."""
        requests = []
        
        # Add sheet
        requests.append({
            'addSheet': {
                'properties': {
                    'sheetId': sheet_id,
                    'title': 'Valuation Models'
                }
            }
        })
        
        # Prepare data rows
        rows = []
        rows.append([section_data['title']])  # Title row
        rows.append([])  # Empty row
        
        for valuation in section_data['valuations']:
            rows.append([valuation['address']])
            rows.append(['Narrative:', valuation['narrative']])
            rows.append(['Valuation Metrics:'])
            
            for metric_name, metric_value in valuation['metrics'].items():
                rows.append([metric_name, metric_value])
            
            rows.append([])  # Empty row between comparables
        
        # Update cells
        requests.append({
            'updateCells': {
                'rows': [{'values': [{'userEnteredValue': {'stringValue': str(cell)}} for cell in row]} for row in rows],
                'fields': 'userEnteredValue',
                'start': {'sheetId': sheet_id, 'rowIndex': 0, 'columnIndex': 0}
            }
        })
        
        return requests
    
    def _create_arv_range_requests(self, section_data: Dict[str, Any], sheet_id: int) -> List[Dict[str, Any]]:
        """Create Google Sheets requests for ARV range section."""
        requests = []
        
        # Add sheet
        requests.append({
            'addSheet': {
                'properties': {
                    'sheetId': sheet_id,
                    'title': 'ARV Range'
                }
            }
        })
        
        # Prepare data rows
        rows = []
        rows.append([section_data['title']])  # Title row
        rows.append([])  # Empty row
        
        for arv_type, arv_value in section_data['arv_range'].items():
            rows.append([arv_type, arv_value])
        
        # Update cells
        requests.append({
            'updateCells': {
                'rows': [{'values': [{'userEnteredValue': {'stringValue': str(cell)}} for cell in row]} for row in rows],
                'fields': 'userEnteredValue',
                'start': {'sheetId': sheet_id, 'rowIndex': 0, 'columnIndex': 0}
            }
        })
        
        return requests
    
    def _create_key_drivers_requests(self, section_data: Dict[str, Any], sheet_id: int) -> List[Dict[str, Any]]:
        """Create Google Sheets requests for key drivers section."""
        requests = []
        
        # Add sheet
        requests.append({
            'addSheet': {
                'properties': {
                    'sheetId': sheet_id,
                    'title': 'Key Drivers'
                }
            }
        })
        
        # Prepare data rows
        rows = []
        rows.append([section_data['title']])  # Title row
        rows.append([])  # Empty row
        
        for idx, driver in enumerate(section_data['drivers'], start=1):
            rows.append([f"{idx}. {driver}"])
        
        # Update cells
        requests.append({
            'updateCells': {
                'rows': [{'values': [{'userEnteredValue': {'stringValue': str(cell)}} for cell in row]} for row in rows],
                'fields': 'userEnteredValue',
                'start': {'sheetId': sheet_id, 'rowIndex': 0, 'columnIndex': 0}
            }
        })
        
        return requests
