"""
Excel export service for multifamily pro forma workbooks.

Produces a 10-sheet .xlsx workbook using openpyxl. The sheet structure is
driven entirely by `excel_workbook_spec.WORKBOOK_SHEETS` so that the export
and import stay in sync.

- Computed-output sheets (S00a, S00b, S05, S06) are populated from the
  ProFormaComputation (values only, no formulas — Req 12.3).
- Round-trippable sheets (S01, S02, S03, S04, S07, Funding_Sources) are
  populated from the Deal's ORM models.

Requirements: 12.1-12.4, 14.5
"""

from __future__ import annotations

import io
import logging
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import Workbook

from app.services.multifamily.excel_workbook_spec import (
    ColumnSpec,
    SheetSpec,
    WORKBOOK_SHEETS,
)

logger = logging.getLogger(__name__)


class ExcelExportService:
    """Exports a multifamily Deal to a .xlsx workbook."""

    def export_deal(self, deal_id: int) -> bytes:
        """Export a Deal to an Excel workbook as bytes.

        Args:
            deal_id: The Deal to export.

        Returns:
            The .xlsx file content as bytes.

        Requirements: 12.1-12.4
        """
        from app.models.deal import Deal
        from app.models.unit import Unit
        from app.models.rent_comp import RentComp
        from app.models.sale_comp import SaleComp
        from app.models.funding_source import FundingSource
        from app.models.deal_lender_selection import DealLenderSelection
        from app.services.multifamily.dashboard_service import DashboardService

        deal = Deal.query.get(deal_id)
        if deal is None:
            raise ValueError(f"Deal {deal_id} not found")

        # Get the computed pro forma result for output sheets
        dashboard_service = DashboardService()
        pro_forma_dict = dashboard_service.get_pro_forma(deal_id)

        # Load related data for round-trippable sheets
        units = Unit.query.filter_by(deal_id=deal_id).order_by(Unit.unit_identifier).all()
        rent_comps = RentComp.query.filter_by(deal_id=deal_id).all()
        sale_comps = SaleComp.query.filter_by(deal_id=deal_id).all()
        funding_sources = FundingSource.query.filter_by(deal_id=deal_id).all()
        lender_selections = (
            DealLenderSelection.query
            .filter_by(deal_id=deal_id)
            .all()
        )

        # Build the workbook
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)

        for sheet_spec in WORKBOOK_SHEETS:
            ws = wb.create_sheet(title=sheet_spec.name)
            self._write_headers(ws, sheet_spec)

            if sheet_spec.name == "S00a_Summary_ScenarioA":
                self._write_summary_sheet(ws, sheet_spec, pro_forma_dict, "a")
            elif sheet_spec.name == "S00b_Summary_ScenarioB":
                self._write_summary_sheet(ws, sheet_spec, pro_forma_dict, "b")
            elif sheet_spec.name == "S01_RentRoll_InPlace":
                self._write_rent_roll(ws, sheet_spec, units)
            elif sheet_spec.name == "S02_MarketRents_Comps":
                self._write_rent_comps(ws, sheet_spec, rent_comps)
            elif sheet_spec.name == "S03_SaleComps_CapRates":
                self._write_sale_comps(ws, sheet_spec, sale_comps)
            elif sheet_spec.name == "S04_Rehab_Timing":
                self._write_rehab_timing(ws, sheet_spec, units)
            elif sheet_spec.name == "S05_ProForma_24mo":
                self._write_pro_forma_schedule(ws, sheet_spec, pro_forma_dict)
            elif sheet_spec.name == "S06_Valuation":
                self._write_valuation(ws, sheet_spec, pro_forma_dict)
            elif sheet_spec.name == "S07_Lender_Assumptions":
                self._write_lender_assumptions(ws, sheet_spec, lender_selections)
            elif sheet_spec.name == "Funding_Sources":
                self._write_funding_sources(ws, sheet_spec, funding_sources)

        # Serialize to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # -----------------------------------------------------------------------
    # Internal helpers
    # -----------------------------------------------------------------------

    def _write_headers(self, ws, sheet_spec: SheetSpec) -> None:
        """Write the header row for a sheet."""
        for col_idx, col_spec in enumerate(sheet_spec.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_spec.header)

    def _write_row(self, ws, row_num: int, sheet_spec: SheetSpec, data: dict[str, Any]) -> None:
        """Write a single data row using the sheet spec's column order."""
        for col_idx, col_spec in enumerate(sheet_spec.columns, start=1):
            value = data.get(col_spec.attr)
            ws.cell(row=row_num, column=col_idx, value=self._serialize_cell(value, col_spec))

    def _serialize_cell(self, value: Any, col_spec: ColumnSpec) -> Any:
        """Convert a Python value to an Excel-friendly cell value."""
        if value is None:
            return None
        if col_spec.kind in ("decimal", "rate"):
            if isinstance(value, Decimal):
                return float(value)
            if isinstance(value, str):
                # Canonical dict serializes Decimal as str
                return float(Decimal(value))
            return float(value)
        if col_spec.kind == "int":
            return int(value)
        if col_spec.kind == "bool":
            if isinstance(value, bool):
                return value
            if isinstance(value, str):
                return value.lower() in ("true", "1", "yes")
            return bool(value)
        if col_spec.kind == "date":
            if isinstance(value, date):
                return value
            if isinstance(value, str) and value:
                return value  # openpyxl handles ISO date strings
            return None
        # str
        return str(value) if value is not None else None

    def _write_summary_sheet(
        self, ws, sheet_spec: SheetSpec, pro_forma_dict: dict, scenario_suffix: str
    ) -> None:
        """Write a summary sheet (S00a or S00b) from the pro forma computation."""
        summary = pro_forma_dict.get("summary", {})
        sources_and_uses = pro_forma_dict.get(f"sources_and_uses_{scenario_suffix}")

        metrics = [
            ("In_Place_NOI", summary.get("in_place_noi")),
            ("Stabilized_NOI", summary.get("stabilized_noi")),
            (f"In_Place_DSCR_{scenario_suffix.upper()}", summary.get(f"in_place_dscr_{scenario_suffix}")),
            (f"Stabilized_DSCR_{scenario_suffix.upper()}", summary.get(f"stabilized_dscr_{scenario_suffix}")),
            (f"Cash_On_Cash_{scenario_suffix.upper()}", summary.get(f"cash_on_cash_{scenario_suffix}")),
        ]

        if sources_and_uses:
            metrics.extend([
                ("Loan_Amount", sources_and_uses.get("loan_amount")),
                ("Initial_Cash_Investment", sources_and_uses.get("initial_cash_investment")),
                ("Total_Uses", sources_and_uses.get("total_uses")),
                ("Total_Sources", sources_and_uses.get("total_sources")),
            ])

        for row_idx, (metric_name, value) in enumerate(metrics, start=2):
            ws.cell(row=row_idx, column=1, value=metric_name)
            if value is not None:
                cell_val = float(Decimal(value)) if isinstance(value, str) else (
                    float(value) if isinstance(value, Decimal) else value
                )
                ws.cell(row=row_idx, column=2, value=cell_val)

    def _write_rent_roll(self, ws, sheet_spec: SheetSpec, units) -> None:
        """Write S01_RentRoll_InPlace from Unit + RentRollEntry models."""
        for row_idx, unit in enumerate(units, start=2):
            rent_entry = unit.rent_roll_entry
            data = {
                "unit_identifier": unit.unit_identifier,
                "unit_type": unit.unit_type,
                "beds": unit.beds,
                "baths": unit.baths,
                "sqft": unit.sqft,
                "occupancy_status": unit.occupancy_status,
                "current_rent": rent_entry.current_rent if rent_entry else None,
                "lease_start_date": rent_entry.lease_start_date if rent_entry else None,
                "lease_end_date": rent_entry.lease_end_date if rent_entry else None,
                "notes": rent_entry.notes if rent_entry else None,
            }
            self._write_row(ws, row_idx, sheet_spec, data)

    def _write_rent_comps(self, ws, sheet_spec: SheetSpec, rent_comps) -> None:
        """Write S02_MarketRents_Comps from RentComp models."""
        for row_idx, comp in enumerate(rent_comps, start=2):
            data = {
                "address": comp.address,
                "neighborhood": comp.neighborhood,
                "unit_type": comp.unit_type,
                "observed_rent": comp.observed_rent,
                "sqft": comp.sqft,
                "rent_per_sqft": comp.rent_per_sqft,
                "observation_date": comp.observation_date,
                "source_url": comp.source_url,
            }
            self._write_row(ws, row_idx, sheet_spec, data)

    def _write_sale_comps(self, ws, sheet_spec: SheetSpec, sale_comps) -> None:
        """Write S03_SaleComps_CapRates from SaleComp models."""
        for row_idx, comp in enumerate(sale_comps, start=2):
            data = {
                "address": comp.address,
                "unit_count": comp.unit_count,
                "status": comp.status,
                "sale_price": comp.sale_price,
                "close_date": comp.close_date,
                "observed_cap_rate": comp.observed_cap_rate,
                "observed_ppu": comp.observed_ppu,
                "distance_miles": comp.distance_miles,
            }
            self._write_row(ws, row_idx, sheet_spec, data)

    def _write_rehab_timing(self, ws, sheet_spec: SheetSpec, units) -> None:
        """Write S04_Rehab_Timing from RehabPlanEntry models."""
        row_idx = 2
        for unit in units:
            rehab = unit.rehab_plan_entry
            if rehab is None:
                continue
            data = {
                "unit_id": unit.unit_identifier,
                "renovate_flag": rehab.renovate_flag,
                "current_rent": rehab.current_rent,
                "suggested_post_reno_rent": rehab.suggested_post_reno_rent,
                "underwritten_post_reno_rent": rehab.underwritten_post_reno_rent,
                "rehab_start_month": rehab.rehab_start_month,
                "downtime_months": rehab.downtime_months,
                "rehab_budget": rehab.rehab_budget,
                "scope_notes": rehab.scope_notes,
            }
            self._write_row(ws, row_idx, sheet_spec, data)
            row_idx += 1

    def _write_pro_forma_schedule(self, ws, sheet_spec: SheetSpec, pro_forma_dict: dict) -> None:
        """Write S05_ProForma_24mo from the monthly schedule."""
        monthly_schedule = pro_forma_dict.get("monthly_schedule", [])
        for row_idx, month_row in enumerate(monthly_schedule, start=2):
            self._write_row(ws, row_idx, sheet_spec, month_row)

    def _write_valuation(self, ws, sheet_spec: SheetSpec, pro_forma_dict: dict) -> None:
        """Write S06_Valuation from the valuation result."""
        valuation = pro_forma_dict.get("valuation")
        if valuation is None:
            return

        rows = [
            {
                "metric": "Valuation_At_Cap_Rate",
                "min_value": valuation.get("valuation_at_cap_rate_min"),
                "median_value": valuation.get("valuation_at_cap_rate_median"),
                "average_value": valuation.get("valuation_at_cap_rate_average"),
                "max_value": valuation.get("valuation_at_cap_rate_max"),
            },
            {
                "metric": "Valuation_At_PPU",
                "min_value": valuation.get("valuation_at_ppu_min"),
                "median_value": valuation.get("valuation_at_ppu_median"),
                "average_value": valuation.get("valuation_at_ppu_average"),
                "max_value": valuation.get("valuation_at_ppu_max"),
            },
            {
                "metric": "Custom_Cap_Rate_Valuation",
                "min_value": valuation.get("valuation_at_custom_cap_rate"),
                "median_value": None,
                "average_value": None,
                "max_value": None,
            },
            {
                "metric": "Price_To_Rent_Ratio",
                "min_value": valuation.get("price_to_rent_ratio"),
                "median_value": None,
                "average_value": None,
                "max_value": None,
            },
        ]

        for row_idx, row_data in enumerate(rows, start=2):
            self._write_row(ws, row_idx, sheet_spec, row_data)

    def _write_lender_assumptions(self, ws, sheet_spec: SheetSpec, lender_selections) -> None:
        """Write S07_Lender_Assumptions from DealLenderSelection + LenderProfile."""
        row_idx = 2
        for selection in lender_selections:
            profile = selection.lender_profile
            data = {
                "scenario": selection.scenario,
                "is_primary": selection.is_primary,
                "company": profile.company,
                "lender_type": profile.lender_type,
                "origination_fee_rate": profile.origination_fee_rate,
                "ltv_total_cost": profile.ltv_total_cost,
                "construction_rate": profile.construction_rate,
                "construction_io_months": profile.construction_io_months,
                "construction_term_months": profile.construction_term_months,
                "perm_rate": profile.perm_rate,
                "perm_amort_years": profile.perm_amort_years,
                "min_interest_or_yield": profile.min_interest_or_yield,
                "max_purchase_ltv": profile.max_purchase_ltv,
                "treasury_5y_rate": profile.treasury_5y_rate,
                "spread_bps": profile.spread_bps,
                "term_years": profile.term_years,
                "amort_years": profile.amort_years,
                "prepay_penalty_description": profile.prepay_penalty_description,
            }
            self._write_row(ws, row_idx, sheet_spec, data)
            row_idx += 1

    def _write_funding_sources(self, ws, sheet_spec: SheetSpec, funding_sources) -> None:
        """Write Funding_Sources sheet from FundingSource models."""
        for row_idx, source in enumerate(funding_sources, start=2):
            data = {
                "source_type": source.source_type,
                "total_available": source.total_available,
                "interest_rate": source.interest_rate,
                "origination_fee_rate": source.origination_fee_rate,
            }
            self._write_row(ws, row_idx, sheet_spec, data)
