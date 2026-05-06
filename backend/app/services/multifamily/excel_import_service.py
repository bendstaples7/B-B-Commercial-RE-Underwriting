"""
Excel import service for multifamily pro forma workbooks.

Parses only the round-trippable sheets (S01, S02, S03, S04, S07,
Funding_Sources) from an uploaded .xlsx workbook and creates a Deal
with all child rows in a single transaction.

Raises UnsupportedImportFormatError for missing sheets (Req 13.2) or
missing required columns (Req 13.3).

Requirements: 13.1-13.4
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app import db
from app.exceptions import UnsupportedImportFormatError
from app.services.multifamily.excel_workbook_spec import (
    ColumnSpec,
    SheetSpec,
    get_round_trippable_sheets,
    get_required_columns,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Result dataclass
# ---------------------------------------------------------------------------


@dataclass
class SheetParseReport:
    """Parse report for a single sheet."""

    sheet_name: str
    rows_parsed: int = 0
    rows_skipped: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportResult:
    """Result of a workbook import operation."""

    deal_id: int
    parse_report: list[SheetParseReport] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Import service
# ---------------------------------------------------------------------------


class ExcelImportService:
    """Imports a multifamily Deal from a .xlsx workbook."""

    def import_workbook(self, user_id: str, file: BinaryIO) -> ImportResult:
        """Import a workbook and create a Deal with child rows.

        Args:
            user_id: The user performing the import.
            file: A file-like object containing the .xlsx bytes.

        Returns:
            ImportResult with the new deal_id and per-sheet parse reports.

        Raises:
            UnsupportedImportFormatError: If a required sheet or column is missing.

        Requirements: 13.1-13.4
        """
        from app.models.deal import Deal
        from app.models.unit import Unit
        from app.models.rent_roll_entry import RentRollEntry
        from app.models.rent_comp import RentComp
        from app.models.sale_comp import SaleComp
        from app.models.rehab_plan_entry import RehabPlanEntry
        from app.models.lender_profile import LenderProfile
        from app.models.deal_lender_selection import DealLenderSelection
        from app.models.funding_source import FundingSource

        # Load the workbook
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        # Validate all round-trippable sheets are present (Req 13.2)
        round_trippable_sheets = get_round_trippable_sheets()
        for sheet_spec in round_trippable_sheets:
            if sheet_spec.name not in sheet_names:
                raise UnsupportedImportFormatError(
                    f"Required sheet '{sheet_spec.name}' is missing from the workbook",
                    missing_sheet=sheet_spec.name,
                )

        # Validate required columns in each round-trippable sheet (Req 13.3)
        for sheet_spec in round_trippable_sheets:
            ws = wb[sheet_spec.name]
            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            required_cols = get_required_columns(sheet_spec)
            for col_spec in required_cols:
                if col_spec.header not in header_row:
                    raise UnsupportedImportFormatError(
                        f"Required column '{col_spec.header}' is missing from sheet '{sheet_spec.name}'",
                        missing_column=col_spec.header,
                        sheet=sheet_spec.name,
                    )

        parse_reports: list[SheetParseReport] = []

        # Parse S01 — Rent Roll (creates Deal + Units + RentRollEntries)
        s01_spec = self._get_spec_by_name(round_trippable_sheets, "S01_RentRoll_InPlace")
        s01_rows, s01_report = self._parse_sheet(wb, s01_spec)
        parse_reports.append(s01_report)

        # Create the Deal with minimal required fields
        # We derive unit_count from the rent roll rows
        unit_count = max(len(s01_rows), 5)  # Minimum 5 per constraint
        deal = Deal(
            created_by_user_id=user_id,
            property_address="Imported Deal",
            unit_count=unit_count,
            purchase_price=Decimal("1000000"),  # Placeholder
            closing_costs=Decimal("0"),
            status="draft",
        )
        db.session.add(deal)
        db.session.flush()  # Get the deal_id

        # Create Units and RentRollEntries from S01
        unit_map: dict[str, Unit] = {}  # unit_identifier -> Unit
        for row_data in s01_rows:
            unit_identifier = row_data.get("unit_identifier", "")
            if not unit_identifier:
                s01_report.rows_skipped += 1
                continue

            unit = Unit(
                deal_id=deal.id,
                unit_identifier=str(unit_identifier),
                unit_type=row_data.get("unit_type"),
                beds=row_data.get("beds"),
                baths=row_data.get("baths"),
                sqft=row_data.get("sqft"),
                occupancy_status=row_data.get("occupancy_status", "Vacant"),
            )
            db.session.add(unit)
            db.session.flush()
            unit_map[str(unit_identifier)] = unit

            # Create RentRollEntry if current_rent is present
            current_rent = row_data.get("current_rent")
            if current_rent is not None:
                rent_entry = RentRollEntry(
                    unit_id=unit.id,
                    current_rent=current_rent,
                    lease_start_date=row_data.get("lease_start_date"),
                    lease_end_date=row_data.get("lease_end_date"),
                    notes=row_data.get("notes"),
                )
                db.session.add(rent_entry)

        # Update deal unit_count to match actual imported units
        if len(unit_map) >= 5:
            deal.unit_count = len(unit_map)

        # Parse S02 — Rent Comps
        s02_spec = self._get_spec_by_name(round_trippable_sheets, "S02_MarketRents_Comps")
        s02_rows, s02_report = self._parse_sheet(wb, s02_spec)
        parse_reports.append(s02_report)

        for row_data in s02_rows:
            address = row_data.get("address")
            if not address:
                s02_report.rows_skipped += 1
                continue

            sqft = row_data.get("sqft", 1)
            observed_rent = row_data.get("observed_rent", Decimal("0"))
            rent_per_sqft = row_data.get("rent_per_sqft")
            if rent_per_sqft is None and sqft and sqft > 0:
                rent_per_sqft = observed_rent / Decimal(str(sqft))

            rent_comp = RentComp(
                deal_id=deal.id,
                address=str(address),
                neighborhood=row_data.get("neighborhood"),
                unit_type=row_data.get("unit_type", ""),
                observed_rent=observed_rent,
                sqft=sqft if sqft else 1,
                rent_per_sqft=rent_per_sqft or Decimal("0"),
                observation_date=row_data.get("observation_date"),
                source_url=row_data.get("source_url"),
            )
            db.session.add(rent_comp)

        # Parse S03 — Sale Comps
        s03_spec = self._get_spec_by_name(round_trippable_sheets, "S03_SaleComps_CapRates")
        s03_rows, s03_report = self._parse_sheet(wb, s03_spec)
        parse_reports.append(s03_report)

        for row_data in s03_rows:
            address = row_data.get("address")
            if not address:
                s03_report.rows_skipped += 1
                continue

            unit_count_val = row_data.get("unit_count", 1)
            sale_price = row_data.get("sale_price", Decimal("0"))
            observed_ppu = row_data.get("observed_ppu")
            if observed_ppu is None and unit_count_val and unit_count_val > 0:
                observed_ppu = sale_price / Decimal(str(unit_count_val))

            sale_comp = SaleComp(
                deal_id=deal.id,
                address=str(address),
                unit_count=unit_count_val if unit_count_val else 1,
                status=row_data.get("status"),
                sale_price=sale_price,
                close_date=row_data.get("close_date"),
                observed_cap_rate=row_data.get("observed_cap_rate", Decimal("0.05")),
                observed_ppu=observed_ppu or Decimal("0"),
                distance_miles=row_data.get("distance_miles"),
            )
            db.session.add(sale_comp)

        # Parse S04 — Rehab Timing
        s04_spec = self._get_spec_by_name(round_trippable_sheets, "S04_Rehab_Timing")
        s04_rows, s04_report = self._parse_sheet(wb, s04_spec)
        parse_reports.append(s04_report)

        for row_data in s04_rows:
            unit_id_str = row_data.get("unit_id")
            if not unit_id_str or str(unit_id_str) not in unit_map:
                s04_report.rows_skipped += 1
                if unit_id_str:
                    s04_report.warnings.append(
                        f"Unit_ID '{unit_id_str}' not found in rent roll"
                    )
                continue

            unit = unit_map[str(unit_id_str)]
            renovate_flag = row_data.get("renovate_flag", False)
            rehab_start_month = row_data.get("rehab_start_month")
            downtime_months = row_data.get("downtime_months")

            # Compute stabilized_month and stabilizes_after_horizon
            stabilized_month = None
            stabilizes_after_horizon = False
            if renovate_flag and rehab_start_month is not None and downtime_months is not None:
                stabilized_month = rehab_start_month + downtime_months
                if stabilized_month > 24:
                    stabilizes_after_horizon = True

            rehab_entry = RehabPlanEntry(
                unit_id=unit.id,
                renovate_flag=renovate_flag,
                current_rent=row_data.get("current_rent"),
                suggested_post_reno_rent=row_data.get("suggested_post_reno_rent"),
                underwritten_post_reno_rent=row_data.get("underwritten_post_reno_rent"),
                rehab_start_month=rehab_start_month if renovate_flag else None,
                downtime_months=downtime_months if renovate_flag else None,
                stabilized_month=stabilized_month,
                rehab_budget=row_data.get("rehab_budget"),
                scope_notes=row_data.get("scope_notes"),
                stabilizes_after_horizon=stabilizes_after_horizon,
            )
            db.session.add(rehab_entry)

        # Parse S07 — Lender Assumptions
        s07_spec = self._get_spec_by_name(round_trippable_sheets, "S07_Lender_Assumptions")
        s07_rows, s07_report = self._parse_sheet(wb, s07_spec)
        parse_reports.append(s07_report)

        for row_data in s07_rows:
            company = row_data.get("company")
            lender_type = row_data.get("lender_type")
            if not company or not lender_type:
                s07_report.rows_skipped += 1
                continue

            # Create the lender profile
            profile = LenderProfile(
                created_by_user_id=user_id,
                company=str(company),
                lender_type=str(lender_type),
                origination_fee_rate=row_data.get("origination_fee_rate", Decimal("0")),
                prepay_penalty_description=row_data.get("prepay_penalty_description"),
                ltv_total_cost=row_data.get("ltv_total_cost"),
                construction_rate=row_data.get("construction_rate"),
                construction_io_months=row_data.get("construction_io_months"),
                construction_term_months=row_data.get("construction_term_months"),
                perm_rate=row_data.get("perm_rate"),
                perm_amort_years=row_data.get("perm_amort_years"),
                min_interest_or_yield=row_data.get("min_interest_or_yield"),
                max_purchase_ltv=row_data.get("max_purchase_ltv"),
                treasury_5y_rate=row_data.get("treasury_5y_rate"),
                spread_bps=row_data.get("spread_bps"),
                term_years=row_data.get("term_years"),
                amort_years=row_data.get("amort_years"),
            )
            db.session.add(profile)
            db.session.flush()

            # Create the deal-lender selection
            scenario = row_data.get("scenario", "A")
            is_primary = row_data.get("is_primary", False)
            selection = DealLenderSelection(
                deal_id=deal.id,
                lender_profile_id=profile.id,
                scenario=str(scenario),
                is_primary=bool(is_primary),
            )
            db.session.add(selection)

        # Parse Funding_Sources
        funding_spec = self._get_spec_by_name(round_trippable_sheets, "Funding_Sources")
        funding_rows, funding_report = self._parse_sheet(wb, funding_spec)
        parse_reports.append(funding_report)

        for row_data in funding_rows:
            source_type = row_data.get("source_type")
            if not source_type:
                funding_report.rows_skipped += 1
                continue

            funding_source = FundingSource(
                deal_id=deal.id,
                source_type=str(source_type),
                total_available=row_data.get("total_available", Decimal("0")),
                interest_rate=row_data.get("interest_rate", Decimal("0")),
                origination_fee_rate=row_data.get("origination_fee_rate", Decimal("0")),
            )
            db.session.add(funding_source)

        wb.close()

        return ImportResult(deal_id=deal.id, parse_report=parse_reports)

    # -----------------------------------------------------------------------
    # Internal helpers
    # -----------------------------------------------------------------------

    def _get_spec_by_name(
        self, sheets: tuple[SheetSpec, ...], name: str
    ) -> SheetSpec:
        """Get a SheetSpec by name from the round-trippable sheets."""
        for s in sheets:
            if s.name == name:
                return s
        raise ValueError(f"Sheet spec '{name}' not found")

    def _parse_sheet(
        self, wb, sheet_spec: SheetSpec
    ) -> tuple[list[dict[str, Any]], SheetParseReport]:
        """Parse a single sheet into a list of row dicts.

        Returns:
            Tuple of (parsed_rows, parse_report).
        """
        report = SheetParseReport(sheet_name=sheet_spec.name)
        ws = wb[sheet_spec.name]

        # Read header row to build column index mapping
        rows_iter = ws.iter_rows()
        header_row_cells = next(rows_iter)
        headers = [cell.value for cell in header_row_cells]

        # Map header -> column index
        header_to_idx: dict[str, int] = {}
        for idx, header in enumerate(headers):
            if header is not None:
                header_to_idx[str(header)] = idx

        # Map column spec -> column index
        col_mapping: list[tuple[ColumnSpec, int | None]] = []
        for col_spec in sheet_spec.columns:
            idx = header_to_idx.get(col_spec.header)
            col_mapping.append((col_spec, idx))

        # Parse data rows
        parsed_rows: list[dict[str, Any]] = []
        for row_cells in rows_iter:
            cell_values = [cell.value for cell in row_cells]

            # Skip completely empty rows
            if all(v is None for v in cell_values):
                continue

            row_data: dict[str, Any] = {}

            for col_spec, col_idx in col_mapping:
                if col_idx is None:
                    row_data[col_spec.attr] = None
                    continue

                raw_value = cell_values[col_idx] if col_idx < len(cell_values) else None
                parsed_value = self._parse_cell(raw_value, col_spec)
                row_data[col_spec.attr] = parsed_value

            # Check if all required fields are None (skip the row)
            required_cols = [cs for cs, _ in col_mapping if cs.required]
            if all(row_data.get(cs.attr) is None for cs in required_cols):
                report.rows_skipped += 1
                continue

            parsed_rows.append(row_data)
            report.rows_parsed += 1

        return parsed_rows, report

    def _parse_cell(self, value: Any, col_spec: ColumnSpec) -> Any:
        """Parse a raw cell value according to the column spec's kind."""
        if value is None:
            return None

        try:
            if col_spec.kind in ("decimal", "rate"):
                if isinstance(value, (int, float)):
                    return Decimal(str(value))
                if isinstance(value, str):
                    return Decimal(value)
                return Decimal(str(value))
            if col_spec.kind == "int":
                if isinstance(value, float):
                    return int(value)
                return int(value)
            if col_spec.kind == "bool":
                if isinstance(value, bool):
                    return value
                if isinstance(value, str):
                    return value.lower() in ("true", "1", "yes")
                if isinstance(value, (int, float)):
                    return bool(value)
                return False
            if col_spec.kind == "date":
                if isinstance(value, datetime):
                    return value.date()
                if isinstance(value, date):
                    return value
                if isinstance(value, str) and value:
                    return date.fromisoformat(value)
                return None
            # str
            return str(value) if value is not None else None
        except (ValueError, InvalidOperation, TypeError):
            return None
